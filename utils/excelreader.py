from openpyxl import load_workbook
class excelreader:
    @staticmethod
    def get_row(filepath,sheetname,rows=None):
        workbook=load_workbook(filepath)
        sheet=workbook[sheetname]
        headers=[cell.value for cell in sheet[1]]
        data=[]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for header, value in zip(headers, row):
                row_dict[header] = "" if value is None else str(value).strip()
            data.append(row_dict)
        
        if not data:
            raise ValueError("excel is empty")
        if rows is None:
            return data
        if isinstance(rows,list):
            return [data[i]for i in rows]
